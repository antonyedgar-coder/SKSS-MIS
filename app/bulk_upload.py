import csv
import io

from openpyxl import Workbook, load_workbook


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("_", " ")


def read_upload_rows(file_storage):
    """Return list of dict rows from uploaded .xlsx or .csv file."""
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".csv"):
        return _read_csv(file_storage)
    if filename.endswith(".xlsx"):
        return _read_xlsx(file_storage)
    raise ValueError("Please upload a .xlsx or .csv file.")


def _read_csv(file_storage):
    raw = file_storage.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw))
    if not reader.fieldnames:
        raise ValueError("CSV file has no header row.")
    rows = []
    for row in reader:
        normalized = {_normalize_header(k): (v or "").strip() for k, v in row.items()}
        rows.append(normalized)
    return rows


def _read_xlsx(file_storage):
    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise ValueError("Excel file is empty.")

    header_keys = [_normalize_header(h) for h in headers]
    rows = []
    for values in rows_iter:
        if not values or all(v is None or str(v).strip() == "" for v in values):
            continue
        row = {}
        for idx, key in enumerate(header_keys):
            if not key:
                continue
            val = values[idx] if idx < len(values) else ""
            row[key] = "" if val is None else str(val).strip()
        rows.append(row)
    wb.close()
    return rows


def build_workbook(headers, sample_rows=None):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    if sample_rows:
        for row in sample_rows:
            ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def get_cell(row, *aliases, default=""):
    for alias in aliases:
        key = _normalize_header(alias)
        if key in row and row[key] not in (None, ""):
            return str(row[key]).strip()
    return default


def parse_float_cell(row, *aliases, default=0.0) -> float:
    value = get_cell(row, *aliases, default="")
    if not value:
        return default
    try:
        return float(value.replace(",", ""))
    except ValueError:
        return default
