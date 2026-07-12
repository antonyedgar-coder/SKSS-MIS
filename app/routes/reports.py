from collections import defaultdict
from datetime import date
from io import BytesIO

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func

from app.extensions import db
from app.models import (
    BankAccount,
    Brand,
    Company,
    ExpenseCategory,
    ExpenseGroup,
    ExpensePayment,
    Supplier,
    SupplierBill,
    SupplierCategory,
    SupplierPayment,
)
from app.utils import (
    approver_required,
    build_cash_report_matrix,
    build_supplier_ledger,
    expense_payment_particulars,
    get_supplier_brand_pairs_in_period,
    parse_date,
    supplier_period_totals,
)
from app.activity_log import log_activity
from app.permissions import permission_required

reports_bp = Blueprint("reports", __name__)

HEADER_FONT = Font(bold=True)


def _default_dates():
    today = date.today()
    month_start = today.replace(day=1)
    return month_start.isoformat(), today.isoformat()


def _sorted_active_accounts():
    accounts = BankAccount.query.filter_by(is_active=True).all()
    return sorted(accounts, key=lambda a: (0 if a.account_type == "cash" else 1, a.name.lower()))


def _parse_account_ids(values):
    ids = []
    for value in values or []:
        try:
            ids.append(int(value))
        except (TypeError, ValueError):
            continue
    return ids


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT


def _add_sheet_title(ws, title, from_date, to_date):
    ws.append([title])
    ws.append([f"Period: {from_date.strftime('%d-%b-%Y')} to {to_date.strftime('%d-%b-%Y')}"])
    ws.append([])


@reports_bp.route("/")
@login_required
def index():
    from_date, to_date = _default_dates()
    return render_template("reports/index.html", from_date=from_date, to_date=to_date)


def _supplier_report_view_mode(supplier_id, brand_id, category_id, rows):
    if supplier_id:
        return "supplier_ledger"
    if brand_id:
        return "brand_detail"
    if category_id:
        return "category_detail"

    brand_ids = {row["brand"].id for row in rows}
    category_names = {
        row["supplier"].category.name if row["supplier"].category else "Uncategorized"
        for row in rows
    }
    if len(brand_ids) > 1:
        return "brand_summary"
    if len(category_names) > 1:
        return "category_summary"
    return "supplier_summary"


def _build_supplier_report_rows(fd, td, supplier_id=None, brand_id=None, category_id=None):
    rows = []
    category_totals = defaultdict(lambda: {"credit": 0.0, "debit": 0.0, "net": 0.0})
    brand_summary = defaultdict(lambda: {"credit": 0.0, "debit": 0.0, "net": 0.0})
    brand_supplier_summary = defaultdict(lambda: {"credit": 0.0, "debit": 0.0, "net": 0.0})
    category_supplier_summary = defaultdict(lambda: {"credit": 0.0, "debit": 0.0, "net": 0.0})

    for sid, bid in get_supplier_brand_pairs_in_period(fd, td):
        if supplier_id and str(sid) != str(supplier_id):
            continue
        if brand_id and str(bid) != str(brand_id):
            continue

        supplier = db.session.get(Supplier, sid)
        brand = db.session.get(Brand, bid)
        if not supplier or not brand:
            continue
        if category_id and (
            not supplier.category_id or str(supplier.category_id) != str(category_id)
        ):
            continue

        credit, debit, net = supplier_period_totals(sid, fd, td, bid)
        if credit == 0 and debit == 0:
            continue

        cat_name = supplier.category.name if supplier.category else "Uncategorized"
        row = {
            "supplier": supplier,
            "brand": brand,
            "category_name": cat_name,
            "credit": credit,
            "debit": debit,
            "net": net,
        }
        rows.append(row)

        category_totals[cat_name]["credit"] += credit
        category_totals[cat_name]["debit"] += debit
        category_totals[cat_name]["net"] += net
        brand_summary[brand.name]["credit"] += credit
        brand_summary[brand.name]["debit"] += debit
        brand_summary[brand.name]["net"] += net
        brand_supplier_summary[(brand.name, supplier.name)]["credit"] += credit
        brand_supplier_summary[(brand.name, supplier.name)]["debit"] += debit
        brand_supplier_summary[(brand.name, supplier.name)]["net"] += net
        category_supplier_summary[(cat_name, supplier.name)]["credit"] += credit
        category_supplier_summary[(cat_name, supplier.name)]["debit"] += debit
        category_supplier_summary[(cat_name, supplier.name)]["net"] += net

    rows.sort(key=lambda r: (r["supplier"].name, r["brand"].name))

    brand_supplier_rows = [
        {
            "brand_name": key[0],
            "supplier_name": key[1],
            "credit": values["credit"],
            "debit": values["debit"],
            "net": values["net"],
        }
        for key, values in sorted(brand_supplier_summary.items())
    ]
    category_supplier_rows = [
        {
            "category_name": key[0],
            "supplier_name": key[1],
            "credit": values["credit"],
            "debit": values["debit"],
            "net": values["net"],
        }
        for key, values in sorted(category_supplier_summary.items())
    ]

    return rows, dict(category_totals), dict(brand_summary), brand_supplier_rows, category_supplier_rows


@reports_bp.route("/supplier-mis", methods=["GET", "POST"])
@login_required
@permission_required("supplier_mis_report", "view")
def supplier_mis():
    from_date, to_date = _default_dates()
    supplier_id = None
    brand_id = None
    category_id = None
    show_report = False

    if request.method == "POST":
        show_report = True
        from_date = request.form.get("from_date", from_date)
        to_date = request.form.get("to_date", to_date)
        supplier_id = request.form.get("supplier_id") or None
        brand_id = request.form.get("brand_id") or None
        category_id = request.form.get("category_id") or None

    fd = parse_date(from_date)
    td = parse_date(to_date)
    suppliers = Supplier.query.order_by(Supplier.name).all()
    brands = Brand.query.join(Company).order_by(Company.name, Brand.name).all()
    categories = SupplierCategory.query.order_by(SupplierCategory.name).all()

    rows = []
    category_totals = {}
    brand_totals_map = {}
    brand_supplier_rows = []
    category_supplier_rows = []
    ledger_rows = []
    view_mode = "supplier_summary"
    selected_supplier = None

    if show_report and fd and td:
        rows, category_totals, brand_totals_map, brand_supplier_rows, category_supplier_rows = (
            _build_supplier_report_rows(fd, td, supplier_id, brand_id, category_id)
        )
        view_mode = _supplier_report_view_mode(supplier_id, brand_id, category_id, rows)
        if supplier_id:
            selected_supplier = db.session.get(Supplier, int(supplier_id))
            bid = int(brand_id) if brand_id else None
            cid = int(category_id) if category_id else None
            ledger_rows = build_supplier_ledger(int(supplier_id), fd, td, bid, cid)

    return render_template(
        "reports/supplier_mis.html",
        from_date=from_date,
        to_date=to_date,
        suppliers=suppliers,
        brands=brands,
        categories=categories,
        supplier_id=supplier_id,
        brand_id=brand_id,
        category_id=category_id,
        show_report=show_report,
        view_mode=view_mode,
        selected_supplier=selected_supplier,
        rows=rows,
        ledger_rows=ledger_rows,
        category_totals=category_totals,
        brand_totals_map=brand_totals_map,
        brand_supplier_rows=brand_supplier_rows,
        category_supplier_rows=category_supplier_rows,
    )


@reports_bp.route("/supplier-mis/excel")
@login_required
@permission_required("supplier_mis_report", "view")
def supplier_mis_excel():
    from_date = parse_date(request.args.get("from_date"))
    to_date = parse_date(request.args.get("to_date"))
    supplier_id = request.args.get("supplier_id")
    brand_id = request.args.get("brand_id")
    category_id = request.args.get("category_id")

    if not from_date or not to_date:
        flash("Select valid dates.", "danger")
        return redirect(url_for("reports.supplier_mis"))

    rows, category_data, brand_data, brand_supplier_rows, category_supplier_rows = (
        _build_supplier_report_rows(from_date, to_date, supplier_id, brand_id, category_id)
    )
    view_mode = _supplier_report_view_mode(supplier_id, brand_id, category_id, rows)

    wb = Workbook()
    wb.remove(wb.active)

    if view_mode == "supplier_ledger" and supplier_id:
        supplier = db.session.get(Supplier, int(supplier_id))
        bid = int(brand_id) if brand_id else None
        cid = int(category_id) if category_id else None
        ledger_rows = build_supplier_ledger(int(supplier_id), from_date, to_date, bid, cid)
        ws = wb.create_sheet("Supplier Ledger")
        title = f"Supplier Ledger - {supplier.name if supplier else supplier_id}"
        _add_sheet_title(ws, title, from_date, to_date)
        ws.append(["Date", "Particulars", "Reference No.", "Debit", "Credit", "Balance"])
        _style_header(ws, row=4)
        for entry in ledger_rows:
            ws.append(
                [
                    entry["date"].strftime("%d-%b-%Y"),
                    entry["particulars"],
                    entry["reference"],
                    entry["debit"] or "",
                    entry["credit"] or "",
                    entry["balance"],
                ]
            )
    elif view_mode == "brand_detail":
        ws = wb.create_sheet("Brand Detail")
        _add_sheet_title(ws, "Supplier MIS - Brand Detail", from_date, to_date)
        ws.append(["Supplier", "Category", "Brand", "Debit", "Credit", "Net"])
        _style_header(ws, row=4)
        for row in rows:
            ws.append(
                [
                    row["supplier"].name,
                    row["category_name"],
                    row["brand"].name,
                    row["debit"],
                    row["credit"],
                    row["net"],
                ]
            )
    elif view_mode == "category_detail":
        ws = wb.create_sheet("Category Detail")
        _add_sheet_title(ws, "Supplier MIS - Category Detail", from_date, to_date)
        ws.append(["Supplier", "Category", "Brand", "Debit", "Credit", "Net"])
        _style_header(ws, row=4)
        for row in rows:
            ws.append(
                [
                    row["supplier"].name,
                    row["category_name"],
                    row["brand"].name,
                    row["debit"],
                    row["credit"],
                    row["net"],
                ]
            )
    elif view_mode == "brand_summary":
        ws = wb.create_sheet("Brand Summary")
        _add_sheet_title(ws, "Supplier MIS - Brand Summary", from_date, to_date)
        ws.append(["Brand", "Supplier", "Debit", "Credit", "Net"])
        _style_header(ws, row=4)
        for row in brand_supplier_rows:
            ws.append(
                [row["brand_name"], row["supplier_name"], row["debit"], row["credit"], row["net"]]
            )
    elif view_mode == "category_summary":
        ws = wb.create_sheet("Category Summary")
        _add_sheet_title(ws, "Supplier MIS - Category Summary", from_date, to_date)
        ws.append(["Category", "Supplier", "Debit", "Credit", "Net"])
        _style_header(ws, row=4)
        for row in category_supplier_rows:
            ws.append(
                [
                    row["category_name"],
                    row["supplier_name"],
                    row["debit"],
                    row["credit"],
                    row["net"],
                ]
            )
    else:
        ws_all = wb.create_sheet("Supplier Summary")
        _add_sheet_title(ws_all, "Supplier MIS - Supplier Summary", from_date, to_date)
        ws_all.append(
            ["Supplier", "Company", "Brand", "Category", "Debit", "Credit", "Net"]
        )
        _style_header(ws_all, row=4)
        for row in rows:
            s = row["supplier"]
            ws_all.append(
                [
                    s.name,
                    row["brand"].company.name if row["brand"].company else "",
                    row["brand"].name,
                    row["category_name"],
                    row["debit"],
                    row["credit"],
                    row["net"],
                ]
            )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Supplier_MIS_{from_date}_{to_date}.xlsx"
    log_activity("export", "supplier_mis_report", f"Exported Supplier MIS report ({from_date} to {to_date})", filename)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _expenses_base_query(fd, td, category_id=None, group_id=None):
    q = (
        ExpensePayment.query.join(ExpenseCategory)
        .filter(ExpensePayment.payment_date >= fd, ExpensePayment.payment_date <= td)
    )
    if category_id:
        q = q.filter(ExpenseCategory.id == int(category_id))
    if group_id:
        q = q.filter(ExpenseCategory.group_id == int(group_id))
    return q


def _expenses_report_view_mode(category_id, group_id, payments):
    if category_id:
        return "category_detail"
    if group_id:
        return "group_detail"
    group_names = {
        p.category.group.name if p.category and p.category.group else "Ungrouped"
        for p in payments
    }
    if len(group_names) > 1:
        return "group_summary"
    return "category_summary"


def _build_expense_category_totals(fd, td, category_id=None, group_id=None):
    q = (
        db.session.query(
            ExpenseCategory.name,
            func.coalesce(func.sum(ExpensePayment.amount), 0),
        )
        .join(ExpensePayment)
        .filter(
            ExpensePayment.payment_date >= fd,
            ExpensePayment.payment_date <= td,
        )
    )
    if category_id:
        q = q.filter(ExpenseCategory.id == int(category_id))
    if group_id:
        q = q.filter(ExpenseCategory.group_id == int(group_id))
    q = q.group_by(ExpenseCategory.name).order_by(ExpenseCategory.name)
    return {name: float(amt) for name, amt in q.all()}


def _build_expense_group_totals(fd, td, category_id=None, group_id=None):
    group_label = func.coalesce(ExpenseGroup.name, "Ungrouped")
    q = (
        db.session.query(
            group_label,
            func.coalesce(func.sum(ExpensePayment.amount), 0),
        )
        .select_from(ExpensePayment)
        .join(ExpenseCategory)
        .outerjoin(ExpenseGroup)
        .filter(
            ExpensePayment.payment_date >= fd,
            ExpensePayment.payment_date <= td,
        )
    )
    if category_id:
        q = q.filter(ExpenseCategory.id == int(category_id))
    if group_id:
        q = q.filter(ExpenseCategory.group_id == int(group_id))
    q = q.group_by(group_label).order_by(group_label)
    return {name: float(amt) for name, amt in q.all()}


@reports_bp.route("/expenses-mis", methods=["GET", "POST"])
@login_required
@permission_required("expenses_mis_report", "view")
def expenses_mis():
    from_date, to_date = _default_dates()
    category_id = None
    group_id = None
    show_report = False

    if request.method == "POST":
        show_report = True
        from_date = request.form.get("from_date", from_date)
        to_date = request.form.get("to_date", to_date)
        category_id = request.form.get("category_id") or None
        group_id = request.form.get("group_id") or None

    fd = parse_date(from_date)
    td = parse_date(to_date)
    categories = ExpenseCategory.query.order_by(ExpenseCategory.name).all()
    groups = ExpenseGroup.query.order_by(ExpenseGroup.name).all()
    payments = []
    category_totals = {}
    group_totals = {}
    total = 0.0
    view_mode = "category_summary"

    if show_report and fd and td:
        payments = (
            _expenses_base_query(fd, td, category_id, group_id)
            .order_by(ExpensePayment.payment_date, ExpensePayment.id)
            .all()
        )
        total = sum(p.amount for p in payments)
        category_totals = _build_expense_category_totals(fd, td, category_id, group_id)
        group_totals = _build_expense_group_totals(fd, td, category_id, group_id)
        view_mode = _expenses_report_view_mode(category_id, group_id, payments)

    return render_template(
        "reports/expenses_mis.html",
        from_date=from_date,
        to_date=to_date,
        categories=categories,
        groups=groups,
        category_id=category_id,
        group_id=group_id,
        show_report=show_report,
        view_mode=view_mode,
        payments=payments,
        category_totals=category_totals,
        group_totals=group_totals,
        total=total,
    )


@reports_bp.route("/expenses-mis/excel")
@login_required
@permission_required("expenses_mis_report", "view")
def expenses_mis_excel():
    from_date = parse_date(request.args.get("from_date"))
    to_date = parse_date(request.args.get("to_date"))
    category_id = request.args.get("category_id")
    group_id = request.args.get("group_id")

    if not from_date or not to_date:
        flash("Select valid dates.", "danger")
        return redirect(url_for("reports.expenses_mis"))

    payments = (
        _expenses_base_query(from_date, to_date, category_id, group_id)
        .order_by(ExpensePayment.payment_date, ExpensePayment.id)
        .all()
    )
    view_mode = _expenses_report_view_mode(category_id, group_id, payments)
    total = sum(p.amount for p in payments)

    wb = Workbook()
    wb.remove(wb.active)

    if view_mode == "category_detail":
        ws = wb.create_sheet("Category Detail")
        _add_sheet_title(ws, "Expenses MIS - Category Detail", from_date, to_date)
        ws.append(["Date", "Payment No.", "Particulars", "Amount"])
        _style_header(ws, row=4)
        for p in payments:
            ws.append(
                [
                    p.payment_date.strftime("%d-%b-%Y"),
                    p.payment_number or "",
                    expense_payment_particulars(p),
                    p.amount,
                ]
            )
    elif view_mode == "group_detail":
        ws = wb.create_sheet("Group Detail")
        _add_sheet_title(ws, "Expenses MIS - Group Detail", from_date, to_date)
        ws.append(["Date", "Expense Category", "Payment No.", "Amount"])
        _style_header(ws, row=4)
        for p in payments:
            ws.append(
                [
                    p.payment_date.strftime("%d-%b-%Y"),
                    p.category.name,
                    p.payment_number or "",
                    p.amount,
                ]
            )
    elif view_mode == "group_summary":
        ws = wb.create_sheet("Group Summary")
        _add_sheet_title(ws, "Expenses MIS - Group Summary", from_date, to_date)
        ws.append(["Expense Group", "Amount"])
        _style_header(ws, row=4)
        group_totals = _build_expense_group_totals(from_date, to_date, category_id, group_id)
        for name, amt in sorted(group_totals.items()):
            ws.append([name, amt])
    else:
        ws = wb.create_sheet("Category Summary")
        _add_sheet_title(ws, "Expenses MIS - Category Summary", from_date, to_date)
        ws.append(["Category", "Amount"])
        _style_header(ws, row=4)
        category_totals = _build_expense_category_totals(from_date, to_date, category_id, group_id)
        for name, amt in sorted(category_totals.items()):
            ws.append([name, amt])

    ws_sum = wb.create_sheet("Summary", 0)
    _add_sheet_title(ws_sum, "Expenses MIS - Summary", from_date, to_date)
    ws_sum.append(["Total Expenses", total])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Expenses_MIS_{from_date}_{to_date}.xlsx"
    log_activity("export", "expenses_mis_report", f"Exported Expenses MIS report ({from_date} to {to_date})", filename)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@reports_bp.route("/cash-report", methods=["GET", "POST"])
@login_required
@permission_required("cash_report", "view")
def cash_report():
    from_date, to_date = _default_dates()
    show_report = False
    report_data = None
    selected_account_ids = []
    accounts = _sorted_active_accounts()

    if request.method == "POST":
        show_report = True
        from_date = request.form.get("from_date", from_date)
        to_date = request.form.get("to_date", to_date)
        selected_account_ids = _parse_account_ids(request.form.getlist("account_ids"))

    fd = parse_date(from_date)
    td = parse_date(to_date)
    filter_ids = selected_account_ids if selected_account_ids else None

    if show_report and fd and td:
        report_data = build_cash_report_matrix(fd, td, filter_ids)

    return render_template(
        "reports/cash_report.html",
        from_date=from_date,
        to_date=to_date,
        accounts=accounts,
        selected_account_ids=selected_account_ids,
        show_report=show_report,
        report_data=report_data,
    )


@reports_bp.route("/cash-report/excel")
@login_required
@permission_required("cash_report", "view")
def cash_report_excel():
    from_date = parse_date(request.args.get("from_date"))
    to_date = parse_date(request.args.get("to_date"))
    selected_account_ids = _parse_account_ids(request.args.getlist("account_ids"))

    if not from_date or not to_date:
        flash("Select valid dates.", "danger")
        return redirect(url_for("reports.cash_report"))

    filter_ids = selected_account_ids if selected_account_ids else None
    report_data = build_cash_report_matrix(from_date, to_date, filter_ids)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Report"
    _add_sheet_title(ws, "Cash Report", from_date, to_date)

    header = ["Particulars"] + [col["name"] for col in report_data["columns"]] + ["Total"]
    ws.append(header)
    _style_header(ws, row=4)

    for section in report_data["sections"]:
        ws.append([section["title"]] + [""] * len(report_data["columns"]) + [""])
        section_row = ws.max_row
        for cell in ws[section_row]:
            cell.font = HEADER_FONT
        for row in section["rows"]:
            ws.append(
                [row["label"]]
                + [row["amounts"].get(col["id"], 0.0) for col in report_data["columns"]]
                + [row["total"]]
            )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Cash_Report_{from_date}_{to_date}.xlsx"
    log_activity("export", "cash_report", f"Exported Cash report ({from_date} to {to_date})", filename)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@reports_bp.route("/po-variation", methods=["GET", "POST"])
@login_required
@permission_required("po_variation_report", "view")
def po_variation_report():
    from_date, to_date = _default_dates()
    status_filter = "with_variation"
    if request.method == "POST":
        from_date = request.form.get("from_date", from_date)
        to_date = request.form.get("to_date", to_date)
        status_filter = request.form.get("status_filter", status_filter)
    else:
        from_date = request.args.get("from_date", from_date)
        to_date = request.args.get("to_date", to_date)
        status_filter = request.args.get("status_filter", status_filter)

    fd = parse_date(from_date)
    td = parse_date(to_date)
    rows = []
    pending_count = 0

    if fd and td:
        q = SupplierBill.query.filter(
            SupplierBill.purchase_order_id.isnot(None),
            SupplierBill.bill_date >= fd,
            SupplierBill.bill_date <= td,
        )
        if status_filter == "pending":
            q = q.filter(SupplierBill.variation_status == "pending")
        elif status_filter == "closed":
            q = q.filter(SupplierBill.variation_status == "closed")
        elif status_filter == "with_variation":
            q = q.filter(SupplierBill.variation_status.in_(["pending", "closed"]))

        bills = q.order_by(SupplierBill.bill_date.desc()).all()
        for bill in bills:
            if status_filter == "all" and not bill.has_variation and bill.variation_status == "none":
                continue
            if bill.variation_status == "pending":
                pending_count += 1
            rows.append(bill)

    return render_template(
        "reports/po_variation.html",
        from_date=from_date,
        to_date=to_date,
        status_filter=status_filter,
        rows=rows,
        pending_count=pending_count,
    )


@reports_bp.route("/po-variation/<int:bill_id>/close", methods=["POST"])
@login_required
@approver_required
def close_po_variation(bill_id):
    from datetime import datetime

    bill = db.get_or_404(SupplierBill, bill_id)
    if bill.variation_status != "pending":
        flash("This variation is not pending approval.", "warning")
    else:
        bill.variation_status = "closed"
        bill.variation_closed_by_id = current_user.id
        bill.variation_closed_at = datetime.utcnow()
        bill.approver_remarks = request.form.get("approver_remarks", "").strip() or None
        db.session.commit()
        log_activity(
            "approve",
            "po_variation_report",
            f"Closed PO variation for bill {bill.bill_no}",
            bill.bill_no,
        )
        flash("Variation closed by approver.", "success")

    from urllib.parse import urlencode

    params = urlencode(
        {
            "from_date": request.form.get("from_date", ""),
            "to_date": request.form.get("to_date", ""),
            "status_filter": request.form.get("status_filter", "with_variation"),
        }
    )
    return redirect(f"{url_for('reports.po_variation_report')}?{params}")


@reports_bp.route("/po-variation/excel")
@login_required
@permission_required("po_variation_report", "view")
def po_variation_excel():
    from_date = parse_date(request.args.get("from_date"))
    to_date = parse_date(request.args.get("to_date"))
    status_filter = request.args.get("status_filter", "all")

    if not from_date or not to_date:
        flash("Select valid dates.", "danger")
        return redirect(url_for("reports.po_variation_report"))

    q = SupplierBill.query.filter(
        SupplierBill.purchase_order_id.isnot(None),
        SupplierBill.bill_date >= from_date,
        SupplierBill.bill_date <= to_date,
    )
    if status_filter == "pending":
        q = q.filter(SupplierBill.variation_status == "pending")
    elif status_filter == "closed":
        q = q.filter(SupplierBill.variation_status == "closed")
    elif status_filter == "with_variation":
        q = q.filter(SupplierBill.variation_status.in_(["pending", "closed"]))

    bills = q.order_by(SupplierBill.bill_date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "PO Variation"
    _add_sheet_title(ws, "Purchase Order Variation Report", from_date, to_date)
    ws.append(
        [
            "Bill Date",
            "Order No.",
            "Supplier",
            "Brand",
            "Ordered Amount",
            "Purchase Amount",
            "Variation",
            "Status",
            "Closed By",
            "Approver Remarks",
        ]
    )
    _style_header(ws, row=4)

    for bill in bills:
        if status_filter == "all" and not bill.has_variation and bill.variation_status == "none":
            continue
        brand_label = f"{bill.brand.company.name} - {bill.brand.name}" if bill.brand else ""
        ws.append(
            [
                bill.bill_date.strftime("%d-%b-%Y"),
                bill.bill_no,
                bill.supplier.name,
                brand_label,
                bill.ordered_amount,
                bill.total_amount,
                bill.variation_amount,
                bill.variation_status,
                bill.variation_closed_by.username if bill.variation_closed_by else "",
                bill.approver_remarks or "",
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"PO_Variation_{from_date}_{to_date}.xlsx"
    log_activity("export", "po_variation_report", f"Exported PO Variation report ({from_date} to {to_date})", filename)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
